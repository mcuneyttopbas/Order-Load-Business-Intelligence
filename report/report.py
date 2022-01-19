import datetime
from PyQt5 import QtWidgets,QtCore,QtGui
from PyQt5.QtWidgets import QTableWidgetItem, QAction, QFileDialog, QMessageBox
from report._report import Ui_report_form
import pymongo
from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import random


class  ReportForm(QtWidgets.QMainWindow):
    def __init__(self):
        super(ReportForm,self).__init__()
        print("Report Form is called.")

    def start(self,username,password):
            self.username = username
            self.password = password

            self.ui = Ui_report_form()
            self.ui.setupUi(self)
            self.setWindowIcon(QtGui.QIcon('icon.png'))

            self.ui.radio_both.setEnabled(False)
            self.ui.radio_drapery.setEnabled(False)
            self.ui.radio_tulle.setEnabled(False)
            self.ui.radio_upholstery.setEnabled(False)
            self.ui.txt_orderCode.setEnabled(False)

            self.connect_to_db()

            self.quit = QAction("Quit",self)
            self.quit.triggered.connect(self.closeEvent)

            today = date.today()
            year = today.year
            month = today.month
            day = today.day
            print(month)
            print(type(month))
            if month == 1:
                self.ui.date_start.setDate(QtCore.QDate(year-1, 12, day))
            else:
                self.ui.date_start.setDate(QtCore.QDate(year, month-1, day))
            self.ui.date_end.setDate(QtCore.QDate(year, month, day))

            self.load_data()
            self.display_table()

            self.ui.date_start.dateChanged.connect(self.display_table)
            self.ui.date_end.dateChanged.connect(self.display_table)
            self.ui.cb_company.currentTextChanged.connect(self.display_table)
            self.ui.txt_orderCode.textChanged.connect(self.display_table)
            self.ui.btnGroup_orderDetails.buttonToggled[QtWidgets.QAbstractButton, bool].connect(self.display_table)
            self.ui.btnGroup_category.buttonToggled[QtWidgets.QAbstractButton, bool].connect(self.display_table)
            self.ui.btnGroup_display.buttonToggled[QtWidgets.QAbstractButton, bool].connect(self.display_table)
            self.ui.btn_convert.clicked.connect(self.convert_excel)

    def warning_messageBox(self,explanation):
        msg = QMessageBox()
        msg.setWindowTitle("Uyarı")
        msg.setText(explanation)
        msg.setIcon(QMessageBox.Warning)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setWindowIcon(QtGui.QIcon('icon.png'))
        msg.raise_()
        x = msg.exec_() 

    def connect_to_db(self):
        try:
            self.myclient = pymongo.MongoClient(f"mongodb+srv://{self.username}:{self.password}@cluster0.asdnj.mongodb.net/app_test?retryWrites=true&w=majority")
            self.mydb = self.myclient["order-load"]
            self.order_coll = self.mydb["orders"]
            self.customer_coll = self.mydb["customers"]
        except pymongo.errors.PyMongoError:
            self.warning_messageBox("İnternet bağlantınızı kontrol ediniz!")
            self.close()

    def load_data(self):
        try:
            #TO LOAD COMPANY COMBOBOX
            self.ui.cb_company.addItem("")
            for customer in self.customer_coll.find():
                customer_name = customer["_id"]
                self.ui.cb_company.addItem(customer_name)
        except pymongo.errors.PyMongoError:
            self.warning_messageBox("İnternet bağlantınızı kontrol ediniz!")
            self.close()

    def update_table(self):
        print("updating")
        
    def display_table(self):
        try:
            self.ui.table_statics.clear()
            meter = 0
            record = 0
            
            ############################ COLUMN SETTINGS #################################
            ##############################################################################
            self.column_info  = {
                "Sipariş Kodu" : {"width": 100, "path":"_id","excel":20},
                "Firma":{"width": 100, "path":"Company Name","excel":20},
                "ID" : {"width":30,"excel":8},
                "Durum":{"width":100, "path":"Status","excel":20},
                "Ürün":{"width":100, "path":"Item","excel":20},
                "Renk Kodu": {"width":100, "path":"Color","excel":12},
                "Metre" :{"width":30, "path":"Meter","excel":8},
                "Alıcı" : {"width":75,"excel":20},
                "Kargo" : {"width":75,"excel":20},
                "Sipariş Tarihi":{"width":100,"path":"Order_at","excel":25},
                "Düzenlenme Tarihi":{"width":150,"path":"Created_at","excel":25},
                "Sevk Tarihi":{"width":100, "path":"Delivered_at","excel":25},
                "Not":{"width":150,"path":"Note","excel":30}
            }

            self.selected_display_items = []
            for button in self.ui.btnGroup_display.buttons():
                if button.isChecked() == True:
                    self.selected_display_items.append(button.text())

            self.ui.table_statics.setColumnCount(len(self.selected_display_items))  
            self.ui.table_statics.setHorizontalHeaderLabels((item for item in self.selected_display_items))
            column_index = 0

            for item in self.selected_display_items:     
                self.ui.table_statics.setColumnWidth(column_index,self.column_info[item]["width"])
                column_index += 1
            ###############################################################################

            self.selected_filter_items = []
            for button in self.ui.btnGroup_orderDetails.buttons():
                if button.isChecked() == True:
                    self.selected_filter_items.append(button.objectName()[:-12])

            rowCount = 0
            for order in self.order_coll.find():
                if self.ui.cb_company.currentText() != "":
                    if self.ui.cb_company.currentText() != order["Company Name"]:
                        continue 
                for order_items in order["Order Details"]:
                    if order['Order Details'][order_items]['Status'] in self.selected_filter_items :
                        create_date = order["Created_at"].split()[0].split("-") 
                        year, month, day = create_date
                        date = datetime.datetime(int(year), int(month), int(day))
                        if self.ui.date_start.date() <= date <= self.ui.date_end.date():
                            rowCount += 1 
            self.ui.table_statics.setRowCount(rowCount)

            row_index = 0
            for order in self.order_coll.find():
                if self.ui.cb_company.currentText() != "":
                    if self.ui.cb_company.currentText() != order["Company Name"]:
                        continue 
                for order_items in order["Order Details"]:
                    create_date = order["Created_at"].split()[0].split("-") 
                    year, month, day = create_date
                    date = datetime.datetime(int(year), int(month), int(day))
                    if self.ui.date_start.date() <= date <= self.ui.date_end.date():
                        if order['Order Details'][order_items]['Status'] in self.selected_filter_items :
                            
                            for item in self.selected_display_items:
                                if item == "Sipariş Kodu" or item == "Firma" or item == "Düzenlenme Tarihi" or item == "Sipariş Tarihi":
                                    self.ui.table_statics.setItem(row_index,self.selected_display_items.index(item),QTableWidgetItem(order[self.column_info[item]["path"]]))
                                elif item == "Durum":
                                    if order['Order Details'][order_items]['Status'] == "new":
                                        self.ui.table_statics.setItem(row_index,self.selected_display_items.index(item),QTableWidgetItem("YENİ"))
                                    elif order['Order Details'][order_items]['Status'] == "preparing":
                                        self.ui.table_statics.setItem(row_index,self.selected_display_items.index(item),QTableWidgetItem("HAZIRLANAN"))
                                    elif order['Order Details'][order_items]['Status'] == "waiting":
                                        self.ui.table_statics.setItem(row_index,self.selected_display_items.index(item),QTableWidgetItem("BEKLEYEN"))
                                    elif order['Order Details'][order_items]['Status'] == "ready":
                                        self.ui.table_statics.setItem(row_index,self.selected_display_items.index(item),QTableWidgetItem("TAMAMLANAN"))
                                    elif order['Order Details'][order_items]['Status'] == "delivered":
                                        self.ui.table_statics.setItem(row_index,self.selected_display_items.index(item),QTableWidgetItem("SEVKEDİLEN"))
                                elif item == "ID":
                                    self.ui.table_statics.setItem(row_index,self.selected_display_items.index(item),QTableWidgetItem(order_items))
                                elif item == "Ürün" or item == "Renk Kodu" or item == "Metre" or item == "Not" or item == "Sevk Tarihi":
                                    self.ui.table_statics.setItem(row_index,self.selected_display_items.index(item),QTableWidgetItem(order['Order Details'][order_items][self.column_info[item]["path"]]))
                                    if item == "Metre":
                                        try:
                                            meter += int(order['Order Details'][order_items][self.column_info[item]["path"]])
                                        except ValueError:
                                            meter += float(order['Order Details'][order_items][self.column_info[item]["path"]])
                                elif item == "Alıcı":
                                    self.ui.table_statics.setItem(row_index,self.selected_display_items.index(item),QTableWidgetItem(order["Receiver Info"]["Receiver Name"]))
                                elif item == "Kargo":
                                    self.ui.table_statics.setItem(row_index,self.selected_display_items.index(item),QTableWidgetItem(order["Shipping Info"]["Shipper Name"]))
        
                            row_index += 1
                            record += 1 
                            self.rowIndex = row_index
            
            self.ui.lbl_amount.setText("Toplam Metre :")
            self.ui.lbl_txt_amount.setText(str(meter)+" MT")
            self.ui.lbl_record.setText("Toplam Kayıt :")
            self.ui.lbl_txt_record.setText(str(record))
        except pymongo.errors.PyMongoError:
            self.warning_messageBox("İnternet bağlantınızı kontrol ediniz!")
            self.close()
    
    def convert_excel(self):
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Order & Load Report"
            headings = self.selected_display_items
            ws.append(headings)

            for col in range(len(self.selected_display_items)):
                ws[get_column_letter(col+1) + "1"].font = Font(bold=True)
                ws.column_dimensions[get_column_letter(col+1)].width = self.column_info[self.selected_display_items[col]]["excel"]

            for row in range(self.rowIndex):
                row_data = []
                for column in range(len(self.selected_display_items)):
                    if self.selected_display_items[column] == "Metre":
                        try:
                            formatted_meter = int(self.ui.table_statics.item(row,column).text())
                            row_data.append(formatted_meter)
                        except ValueError:
                            formatted_meter = float(self.ui.table_statics.item(row,column).text())
                            row_data.append(formatted_meter)
                    else:
                        row_data.append(self.ui.table_statics.item(row,column).text())
                ws.append(row_data)

            filecode = random.randint(1,10000000000)
            dir_path = QFileDialog.getExistingDirectory()
            print(type(dir_path))
            print(dir_path)
            if dir_path != "":
                wb.save(dir_path + f"/O&L-{filecode}.xlsx")

                msg = QMessageBox()
                msg.setWindowTitle("İşlem Raporu")
                msg.setText(f" Dosya {f'/O&L-{filecode}.xlsx'} oluşturuldu.")
                msg.setIcon(QMessageBox.Information)
                msg.setStandardButtons(QMessageBox.Ok)
                msg.setWindowIcon(QtGui.QIcon('icon.png'))
                msg.raise_()
                x = msg.exec_()

        except Exception as ex :
            msg = QMessageBox()
            msg.setWindowTitle("Uyarı")
            msg.setText(f"Hata: {ex}")
            msg.setIcon(QMessageBox.Warning)
            msg.setStandardButtons(QMessageBox.Ok)
            msg.setWindowIcon(QtGui.QIcon('icon.png'))
            msg.raise_()
            x = msg.exec_() 

    def closeEvent(self,event):
        print("Close evetn of the Report Form is activated.")
        try:
            self.ui.table_statics.clear()
            self.ui.lbl_txt_record.clear()
            self.ui.lbl_record.clear()
            self.ui.lbl_txt_amount.clear()
            self.ui.lbl_amount.clear()
        except Exception as ex:
            print(f"Close event is activated. Raised : {ex}")
       




